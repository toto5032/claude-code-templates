import sqlite3
import os
from datetime import datetime


class TemplateDB:
    def __init__(self, db_path=None):
        if db_path is None:
            db_path = os.path.join(os.path.dirname(__file__), 'templates.db')
        self.db_path = db_path
        self._init_db()

    def _get_conn(self):
        conn = sqlite3.connect(self.db_path, check_same_thread=False)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        return conn

    def _init_db(self):
        conn = self._get_conn()
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                type TEXT DEFAULT 'new',
                created_at DATETIME DEFAULT (datetime('now', 'localtime')),
                updated_at DATETIME DEFAULT (datetime('now', 'localtime')),
                project_name TEXT,
                tech_stack TEXT,
                work_files TEXT,
                goal TEXT,
                background TEXT,
                requirements TEXT,
                input_desc TEXT,
                output_desc TEXT,
                restrictions TEXT,
                ref_patterns TEXT
            );
        ''')
        conn.commit()
        conn.close()

    def get_all(self):
        conn = self._get_conn()
        rows = conn.execute('SELECT * FROM templates ORDER BY updated_at DESC').fetchall()
        conn.close()
        return [dict(r) for r in rows]

    def get(self, template_id):
        conn = self._get_conn()
        row = conn.execute('SELECT * FROM templates WHERE id = ?', (template_id,)).fetchone()
        conn.close()
        return dict(row) if row else None

    def create(self, data):
        conn = self._get_conn()
        cur = conn.execute(
            '''INSERT INTO templates (name, type, project_name, tech_stack, work_files,
               goal, background, requirements, input_desc, output_desc, restrictions, ref_patterns)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (data.get('name', ''), data.get('type', 'new'), data.get('project_name', ''),
             data.get('tech_stack', ''), data.get('work_files', ''),
             data.get('goal', ''), data.get('background', ''), data.get('requirements', ''),
             data.get('input_desc', ''), data.get('output_desc', ''),
             data.get('restrictions', ''), data.get('ref_patterns', ''))
        )
        conn.commit()
        tid = cur.lastrowid
        conn.close()
        return tid

    def update(self, template_id, data):
        conn = self._get_conn()
        conn.execute(
            '''UPDATE templates SET name=?, type=?, project_name=?, tech_stack=?, work_files=?,
               goal=?, background=?, requirements=?, input_desc=?, output_desc=?,
               restrictions=?, ref_patterns=?, updated_at=datetime('now','localtime')
               WHERE id=?''',
            (data.get('name', ''), data.get('type', 'new'), data.get('project_name', ''),
             data.get('tech_stack', ''), data.get('work_files', ''),
             data.get('goal', ''), data.get('background', ''), data.get('requirements', ''),
             data.get('input_desc', ''), data.get('output_desc', ''),
             data.get('restrictions', ''), data.get('ref_patterns', ''), template_id)
        )
        conn.commit()
        conn.close()

    def delete(self, template_id):
        conn = self._get_conn()
        conn.execute('DELETE FROM templates WHERE id = ?', (template_id,))
        conn.commit()
        conn.close()

    def search(self, keyword):
        conn = self._get_conn()
        like = f'%{keyword}%'
        rows = conn.execute(
            '''SELECT * FROM templates
               WHERE name LIKE ? OR project_name LIKE ? OR tech_stack LIKE ? OR goal LIKE ?
               ORDER BY updated_at DESC''',
            (like, like, like, like)
        ).fetchall()
        conn.close()
        return [dict(r) for r in rows]

    def to_prompt(self, template_id):
        t = self.get(template_id)
        if not t:
            return None
        lines = [f"# {t['name']}"]
        if t['project_name']:
            lines.append(f"\n## 프로젝트명\n{t['project_name']}")
        if t['tech_stack']:
            lines.append(f"\n## 기술스택\n{t['tech_stack']}")
        if t['goal']:
            lines.append(f"\n## 목표\n{t['goal']}")
        if t['background']:
            lines.append(f"\n## 배경\n{t['background']}")
        if t['requirements']:
            lines.append(f"\n## 요구사항\n{t['requirements']}")
        if t['input_desc']:
            lines.append(f"\n## 입력\n{t['input_desc']}")
        if t['output_desc']:
            lines.append(f"\n## 산출물\n{t['output_desc']}")
        if t['restrictions']:
            lines.append(f"\n## 금지사항\n{t['restrictions']}")
        if t['ref_patterns']:
            lines.append(f"\n## 참고패턴\n{t['ref_patterns']}")
        if t['work_files']:
            lines.append(f"\n## 작업파일\n{t['work_files']}")
        return '\n'.join(lines)

    def to_claude_md(self, template_id):
        t = self.get(template_id)
        if not t:
            return None
        lines = [f"# {t['name']}", ""]
        if t['goal']:
            lines.append(f"이 프로젝트의 목표: {t['goal']}")
            lines.append("")
        if t['tech_stack']:
            lines.append(f"## 기술스택\n{t['tech_stack']}\n")
        if t['requirements']:
            lines.append(f"## 요구사항\n{t['requirements']}\n")
        if t['restrictions']:
            lines.append(f"## 금지사항\n{t['restrictions']}\n")
        if t['ref_patterns']:
            lines.append(f"## 참고패턴\n{t['ref_patterns']}\n")
        return '\n'.join(lines)

    def import_from_excel(self, file_path):
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        col_map = {
            '이름': 'name', '유형': 'type', '프로젝트명': 'project_name',
            '기술스택': 'tech_stack', '작업파일': 'work_files',
            '목표': 'goal', '배경': 'background', '요구사항': 'requirements',
            '입력': 'input_desc', '산출물': 'output_desc',
            '금지사항': 'restrictions', '참고패턴': 'ref_patterns'
        }
        imported = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = {}
            for i, val in enumerate(row):
                if i < len(headers) and headers[i] in col_map:
                    data[col_map[headers[i]]] = val or ''
            if data.get('name'):
                self.create(data)
                imported += 1
        wb.close()
        return imported
